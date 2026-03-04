from __future__ import annotations

from datetime import date, datetime
from typing import TYPE_CHECKING, Any

import structlog
from openpyxl import load_workbook

if TYPE_CHECKING:
    from openpyxl.worksheet.worksheet import Worksheet

logger = structlog.get_logger(__name__)


class XlsxParser:
    """Extract content from XLSX spreadsheets.

    Provides two extraction modes:

    * ``extract_content`` -- returns a human-readable text representation
      suitable for LLM consumption.
    * ``extract_structured`` -- returns raw rows as a list of dicts keyed
      by column header, useful for deterministic parsing of well-known
      spreadsheet layouts (e.g. project plan templates).
    """

    def extract_content(self, file_path: str) -> str:
        """Open an XLSX file and return all sheets as structured text.

        Each sheet is rendered as a section with the sheet title, followed
        by the header row and each data row on its own line with column
        headers prepended for clarity.

        Args:
            file_path: Absolute path to the .xlsx file on disk.

        Returns:
            Combined text representation of every sheet in the workbook.
        """
        logger.info("xlsx_parser.extract_start", file_path=file_path)

        wb = load_workbook(file_path, data_only=True)
        sections: list[str] = []

        for sheet_name in wb.sheetnames:
            ws: Worksheet = wb[sheet_name]
            sheet_text = self._extract_sheet_text(ws, sheet_name)
            if sheet_text:
                sections.append(sheet_text)

        wb.close()

        content = "\n\n".join(sections)
        logger.info(
            "xlsx_parser.extract_complete",
            file_path=file_path,
            sheet_count=len(wb.sheetnames),
            content_length=len(content),
        )
        return content

    def extract_structured(self, file_path: str) -> list[dict[str, Any]]:
        """Open an XLSX file and return rows as a list of dicts.

        The first non-empty row of each sheet is treated as the header row.
        Subsequent rows are returned as dictionaries mapping column header
        to cell value.  Rows where every cell is ``None`` are skipped.

        Args:
            file_path: Absolute path to the .xlsx file on disk.

        Returns:
            List of dicts with one entry per data row across all sheets.
        """
        logger.info("xlsx_parser.extract_structured_start", file_path=file_path)

        wb = load_workbook(file_path, data_only=True)
        all_rows: list[dict[str, Any]] = []

        for sheet_name in wb.sheetnames:
            ws: Worksheet = wb[sheet_name]
            rows = self._extract_sheet_rows(ws, sheet_name)
            all_rows.extend(rows)

        wb.close()

        logger.info(
            "xlsx_parser.extract_structured_complete",
            file_path=file_path,
            total_rows=len(all_rows),
        )
        return all_rows

    # ------------------------------------------------------------------
    # Private helpers
    # ------------------------------------------------------------------

    def _extract_sheet_text(self, ws: Worksheet, sheet_name: str) -> str:
        """Convert a single worksheet to structured text."""
        rows_iter = ws.iter_rows(values_only=True)
        headers: list[str] = []

        # Find the first non-empty row to use as headers.
        for row in rows_iter:
            candidate = [self._cell_to_str(c) for c in row]
            if any(candidate):
                headers = candidate
                break

        if not headers:
            return ""

        lines: list[str] = [f"## Sheet: {sheet_name}"]
        lines.append("| " + " | ".join(headers) + " |")
        lines.append("| " + " | ".join("---" for _ in headers) + " |")

        for row in rows_iter:
            values = [self._cell_to_str(c) for c in row]
            if not any(values):
                continue
            lines.append("| " + " | ".join(values) + " |")

        # Only return if there is at least one data row beyond the header.
        if len(lines) <= 3:  # noqa: PLR2004 (sheet name + header + separator)
            return ""

        return "\n".join(lines)

    def _extract_sheet_rows(self, ws: Worksheet, sheet_name: str) -> list[dict[str, Any]]:
        """Convert a single worksheet to a list of header-keyed dicts."""
        rows_iter = ws.iter_rows(values_only=True)
        headers: list[str] = []

        for row in rows_iter:
            candidate = [self._cell_to_str(c) for c in row]
            if any(candidate):
                headers = candidate
                break

        if not headers:
            return []

        result: list[dict[str, Any]] = []
        for row in rows_iter:
            values = list(row)
            if not any(v is not None for v in values):
                continue

            record: dict[str, Any] = {"_sheet": sheet_name}
            for idx, header in enumerate(headers):
                raw_value = values[idx] if idx < len(values) else None
                record[header] = raw_value
            result.append(record)

        return result

    @staticmethod
    def _cell_to_str(value: Any) -> str:
        """Convert a cell value to its string representation.

        Handles ``None``, ``datetime``, ``date``, ``float``, and generic
        objects gracefully so that downstream text never contains Python
        repr noise like ``None``.
        """
        if value is None:
            return ""
        if isinstance(value, datetime):
            return value.strftime("%Y-%m-%d")
        if isinstance(value, date):
            return value.isoformat()
        if isinstance(value, float):
            # Render whole-number floats without the trailing ".0".
            if value == int(value):
                return str(int(value))
            return str(value)
        return str(value).strip()
